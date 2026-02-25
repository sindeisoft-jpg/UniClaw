#!/usr/bin/env python3
"""
Generate a styled Excel quotation (报价单) from JSON input.
Requires: openpyxl (pip install openpyxl).

Usage:
  generate_quotation.py --json <path> --out <path>
  generate_quotation.py --out <path>   (read JSON from stdin)

JSON shape:
  {
    "title": "报价单标题",
    "company": "公司名称",
    "date": "2025-02-23",
    "items": [{"name","spec","qty","unit","unitPrice","amount"}],
    "notes": "可选备注"
  }
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Error: openpyxl is required. Run: pip install openpyxl")

# Style definitions
THIN = Side(border_style="thin", color="303030")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

COLUMNS = ["序号", "品名", "规格型号", "数量", "单位", "单价", "金额"]


def load_input(json_path: str | None) -> dict:
    if json_path:
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return json.load(sys.stdin)


def normalize_items(items: list) -> list:
    out = []
    for i, row in enumerate(items, start=1):
        qty = float(row.get("qty", 0))
        unit_price = float(row.get("unitPrice", 0))
        amount = round(qty * unit_price, 2)
        out.append({
            "index": i,
            "name": str(row.get("name", "")),
            "spec": str(row.get("spec", "")),
            "qty": qty,
            "unit": str(row.get("unit", "")),
            "unitPrice": unit_price,
            "amount": amount,
        })
    return out


def write_quotation(data: dict, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "报价单"

    title = data.get("title", "报价单")
    company = data.get("company", "")
    date_str = data.get("date", "")
    items = normalize_items(data.get("items", []))
    notes = data.get("notes", "")

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    row += 2

    # Company & date
    ws.cell(row=row, column=1, value="公司/卖方：")
    ws.cell(row=row, column=2, value=company)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = BORDER
    row += 1
    ws.cell(row=row, column=1, value="日期：")
    ws.cell(row=row, column=2, value=date_str)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = BORDER
    row += 2

    # Table header
    for col, label in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    row += 1

    # Data rows
    for item in items:
        ws.cell(row=row, column=1, value=item["index"]).alignment = CENTER
        ws.cell(row=row, column=2, value=item["name"]).alignment = LEFT
        ws.cell(row=row, column=3, value=item["spec"]).alignment = LEFT
        ws.cell(row=row, column=4, value=item["qty"]).alignment = CENTER
        ws.cell(row=row, column=5, value=item["unit"]).alignment = CENTER
        ws.cell(row=row, column=6, value=item["unitPrice"]).alignment = RIGHT
        ws.cell(row=row, column=7, value=item["amount"]).alignment = RIGHT
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = BORDER
        row += 1

    # Total row
    total = sum(item["amount"] for item in items)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="合计").font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.cell(row=row, column=7, value=round(total, 2)).font = HEADER_FONT
    ws.cell(row=row, column=7).alignment = RIGHT
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = BORDER
    row += 2

    # Notes
    if notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        ws.cell(row=row, column=1, value=f"备注：{notes}").alignment = LEFT
        ws.cell(row=row, column=1).border = BORDER
        row += 1

    # Column widths
    widths = [6, 18, 16, 8, 6, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Excel quotation from JSON")
    parser.add_argument("--json", default=None, help="Path to JSON file (default: read from stdin)")
    parser.add_argument("--out", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    data = load_input(args.json)
    write_quotation(data, Path(args.out))
    print(f"Generated: {args.out}", file=sys.stderr)


if __name__ == "__main__":
    main()
